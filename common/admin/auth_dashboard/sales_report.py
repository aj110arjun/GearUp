import csv
import io
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, F, ExpressionWrapper, FloatField
from django.db.models.functions import TruncDate
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

from common.orders.models import Order  # Adjust import path as needed


class SalesReportGenerator:
    """Generate sales reports in PDF, Excel, and CSV formats"""
    
    def __init__(self, start_date=None, end_date=None, report_type='detailed'):
        self.start_date = start_date or (timezone.now() - timedelta(days=30))
        self.end_date = end_date or timezone.now()
        self.report_type = report_type
    
    def get_report_data(self):
        """Get data for the report"""
        
        # Base queryset - only completed/delivered orders
        orders = Order.objects.filter(
            created_at__range=[self.start_date, self.end_date],
            order_status='delivered'
        ).select_related('user', 'product', 'product__category', 'variant')
        
        # Summary statistics
        total_orders = orders.count()
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        avg_order_value = total_revenue / total_orders if total_orders > 0 else Decimal('0.00')
        
        # Get all orders in period for status breakdown
        all_orders = Order.objects.filter(
            created_at__range=[self.start_date, self.end_date]
        )
        
        # Product-wise sales
        product_sales = orders.values(
            'product__name',
            'product__brand',
            'product__category__name'
        ).annotate(
            quantity_sold=Sum('quantity'),
            total_revenue=Sum(F('unit_price') * F('quantity'))
        ).order_by('-total_revenue')[:20]
        
        # Category-wise sales
        category_sales = orders.values(
            'product__category__name'
        ).annotate(
            quantity_sold=Sum('quantity'),
            total_revenue=Sum(F('unit_price') * F('quantity'))
        ).order_by('-total_revenue')
        
        # Daily sales trend
        daily_sales = orders.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            order_count=Count('order_id'),
            total_revenue=Sum('total_amount'),
            avg_order_value=ExpressionWrapper(
                Sum('total_amount') / Count('order_id'),
                output_field=FloatField()
            )
        ).order_by('date')
        
        # Payment method breakdown
        payment_breakdown = orders.values('payment_method').annotate(
            order_count=Count('order_id'),
            total_revenue=Sum('total_amount')
        ).order_by('-total_revenue')
        
        return {
            'period': {
                'start_date': self.start_date,
                'end_date': self.end_date,
                'days': (self.end_date - self.start_date).days + 1
            },
            'summary': {
                'total_orders': total_orders,
                'total_revenue': total_revenue,
                'avg_order_value': avg_order_value,
                'orders_completed': orders.filter(order_status='delivered').count(),
                'orders_pending': all_orders.filter(
                    order_status__in=['pending', 'confirmed', 'processing', 'shipped']
                ).count(),
                'total_customers': all_orders.values('user').distinct().count(),
            },
            'product_sales': list(product_sales),
            'category_sales': list(category_sales),
            'daily_sales': list(daily_sales),
            'payment_breakdown': list(payment_breakdown),
            'orders': list(orders.values(
                'order_number',
                'created_at',
                'user__first_name',
                'user__last_name',
                'user__email',
                'product__name',
                'variant__sku',
                'quantity',
                'unit_price',
                'total_amount',
                'order_status',
                'payment_status',
                'payment_method'
            )[:1000])  # Increased limit for custom reports
        }
    
    def generate_pdf_report(self):
        """Generate PDF report using ReportLab with improved styling and pagination"""
        data = self.get_report_data()
        
        # Create buffer for PDF
        buffer = BytesIO()
        
        # Create document with landscape orientation for better table display
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=50,
            title=f"Sales Report - {self.start_date.strftime('%d/%m/%Y')}"
        )
        
        # Create story (content)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles for premium look
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor('#111827'),
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=12,
            textColor=colors.HexColor('#1F2937'),
            fontName='Helvetica-Bold',
            borderPadding=5,
            borderWidth=0,
            borderStyle=None
        )

        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.HexColor('#4B5563'),
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=8,
            textColor=colors.HexColor('#374151'),
            leading=14
        )
        
        # Header - Title and Metadata
        story.append(Paragraph(f"SALES PERFORMANCE REPORT", title_style))
        
        period_info = f"""<para alignment='center' spaceAfter=30>
        <font size=11 color='#6B7280'>
        <b>Period:</b> {self.start_date.strftime('%d/%m/%Y')} to {self.end_date.strftime('%d/%m/%Y')} | 
        <b>Generated:</b> {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}
        </font></para>"""
        story.append(Paragraph(period_info, styles['Normal']))
        story.append(Spacer(1, 10))
        
        # 1. Executive Summary Table
        story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        
        summary_data = [
            [Paragraph("<b>Metric</b>", normal_style), Paragraph("<b>Value</b>", normal_style)],
            ['Total Orders Received', str(data['summary']['total_orders'])],
            ['Total Net Revenue', f"Rs. {data['summary']['total_revenue']:,.2f}"],
            ['Average Order Value (AOV)', f"Rs. {data['summary']['avg_order_value']:,.2f}"],
            ['Delivered Orders', str(data['summary']['orders_completed'])],
            ['Fulfillment Pipeline (Pending/Processing)', str(data['summary']['orders_pending'])],
            ['Unique Customers Served', str(data['summary']['total_customers'])],
            ['Total Analytics Period', f"{data['period']['days']} Days"],
        ]
        
        summary_table = Table(summary_data, colWidths=[240, 240])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFAFA')]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # 2. Top Performing Products
        if data['product_sales']:
            story.append(Paragraph("TOP PERFORMING PRODUCTS", heading_style))
            
            product_data = [[
                Paragraph("<b>#</b>", normal_style), 
                Paragraph("<b>Product Name</b>", normal_style), 
                Paragraph("<b>Brand</b>", normal_style), 
                Paragraph("<b>Category</b>", normal_style), 
                Paragraph("<b>Quantity</b>", normal_style), 
                Paragraph("<b>Total Revenue</b>", normal_style)
            ]]
            
            for idx, product in enumerate(data['product_sales'][:20], 1):
                product_data.append([
                    str(idx),
                    Paragraph(product['product__name'] or 'Unknown', normal_style),
                    product['product__brand'] or '-',
                    product['product__category__name'] or 'Uncategorized',
                    str(product['quantity_sold']),
                    f"Rs.{float(product['total_revenue'] or 0):,.2f}"
                ])
            
            product_table = Table(product_data, colWidths=[30, 220, 90, 100, 60, 110], repeatRows=1)
            product_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (4, 0), (4, -1), 'CENTER'),
                ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(product_table)
            story.append(Spacer(1, 30))

        # 3. Order Details Section (Comprehensive)
        if self.report_type == 'detailed' and data['orders']:
            story.append(Paragraph("DETAILED TRANSACTION LOG", heading_style))
            story.append(Paragraph(f"Listing latest {len(data['orders'])} eligible transactions", subheading_style))
            
            order_data = [[
                Paragraph("<b>Order #</b>", normal_style), 
                Paragraph("<b>Date</b>", normal_style), 
                Paragraph("<b>Customer Details</b>", normal_style), 
                Paragraph("<b>Product</b>", normal_style), 
                Paragraph("<b>Qty</b>", normal_style), 
                Paragraph("<b>Method</b>", normal_style),
                Paragraph("<b>Net Total</b>", normal_style)
            ]]
            
            for order in data['orders']:
                customer_name = f"{order['user__first_name'] or ''} {order['user__last_name'] or ''}".strip()
                if not customer_name:
                    customer_name = order['user__email'].split('@')[0]
                
                method_display = dict(Order.PAYMENT_METHOD_CHOICES).get(order['payment_method'], order['payment_method'])
                
                order_data.append([
                    f"#{order['order_number'][-8:]}",
                    order['created_at'].strftime('%d/%m/%Y') if order['created_at'] else '-',
                    Paragraph(f"<b>{customer_name}</b><br/><font size=8>{order['user__email']}</font>", normal_style),
                    Paragraph(f"{order['product__name'] or 'Unknown'}<br/><font size=8 color='#6B7280'>SKU: {order['variant__sku'] or ''}</font>", normal_style),
                    str(order['quantity']),
                    method_display,
                    f"Rs.{float(order['total_amount'] or 0):,.2f}"
                ])
            
            # Using broader widths for detailed logs
            order_table = Table(order_data, colWidths=[70, 80, 140, 180, 40, 70, 90], repeatRows=1)
            order_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4B5563')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (1, -1), 'CENTER'),
                ('ALIGN', (4, 0), (4, -1), 'CENTER'),
                ('ALIGN', (6, 0), (6, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            story.append(order_table)
            story.append(Spacer(1, 20))
            
            # Grand Total Summary Row (Finalizer)
            grand_total_data = [
                ['', '', '', '', '', 'GRAND TOTAL:', f"Rs. {data['summary']['total_revenue']:,.2f}"]
            ]
            grand_total_table = Table(grand_total_data, colWidths=[70, 80, 140, 180, 40, 70, 90])
            grand_total_table.setStyle(TableStyle([
                ('ALIGN', (5, 0), (6, 0), 'RIGHT'),
                ('FONTNAME', (5, 0), (6, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (5, 0), (6, 0), 10),
                ('TEXTCOLOR', (6, 0), (6, 0), colors.HexColor('#111827')),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
            ]))
            story.append(grand_total_table)
        
        # 4. Category and Method Analysis (Two-column layout if space permits)
        if data['category_sales']:
            story.append(Paragraph("CATEGORY PERFORMANCE", heading_style))
            
            category_data = [[
                Paragraph("<b>Category</b>", normal_style), 
                Paragraph("<b>Units Sold</b>", normal_style), 
                Paragraph("<b>Revenue</b>", normal_style)
            ]]
            for category in data['category_sales']:
                category_data.append([
                    category['product__category__name'] or 'Uncategorized',
                    str(category['quantity_sold']),
                    f"Rs. {float(category['total_revenue'] or 0):,.2f}"
                ])
            
            category_table = Table(category_data, colWidths=[300, 100, 120])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E7FF')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            story.append(category_table)

        # Build footer with page numbers
        def add_header_footer(canvas, doc):
            canvas.saveState()
            
            # Footer
            footer_text = f"GearUp Sales Report | Generated: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Page {doc.page}"
            canvas.setFont('Helvetica', 8)
            canvas.setStrokeColor(colors.lightgrey)
            canvas.line(30, 40, doc.width + 30, 40)
            canvas.drawCentredString(doc.width/2 + 30, 25, footer_text)
            
            # Header Branding (Small logo/text at top right)
            canvas.setFont('Helvetica-Bold', 10)
            canvas.drawRightString(doc.width + 30, doc.height + 70, "GEARUP E-COMMERCE")
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(doc.width + 30, doc.height + 60, "Official Internal Sales Document")
            
            canvas.restoreState()

        try:
            doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
            pdf = buffer.getvalue()
            buffer.close()
            return pdf
        except Exception as e:
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=landscape(letter))
            c.drawString(50, 500, f"System Error Encountered during PDF Generation: {str(e)}")
            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            return pdf
    
    def generate_excel_report(self):
        """Generate Excel report with beautiful styling and multiple sheets"""
        data = self.get_report_data()
        
        # Create Excel writer
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            workbook = writer.book
            
            # ===== STYLES =====
            # Title style
            title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            title_alignment = Alignment(horizontal='center', vertical='center')
            title_style = NamedStyle(name="title_style")
            title_style.font = title_font
            title_style.fill = title_fill
            title_style.alignment = title_alignment
            
            # Header style
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_border = Border(
                left=Side(style='thin', color='1F4E78'),
                right=Side(style='thin', color='1F4E78'),
                top=Side(style='thin', color='1F4E78'),
                bottom=Side(style='thin', color='1F4E78')
            )
            header_style = NamedStyle(name="header_style")
            header_style.font = header_font
            header_style.fill = header_fill
            header_style.alignment = header_alignment
            header_style.border = header_border
            
            # Metric style (for summary)
            metric_font = Font(name='Calibri', size=10, bold=True, color='2C3E50')
            metric_style = NamedStyle(name="metric_style")
            metric_style.font = metric_font
            
            # Value style (for summary)
            value_font = Font(name='Calibri', size=10, color='2C3E50')
            value_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            value_alignment = Alignment(horizontal='right')
            value_style = NamedStyle(name="value_style")
            value_style.font = value_font
            value_style.fill = value_fill
            value_style.alignment = value_alignment
            
            # Data style
            data_font = Font(name='Calibri', size=10)
            data_alignment = Alignment(vertical='center')
            data_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            data_style = NamedStyle(name="data_style")
            data_style.font = data_font
            data_style.alignment = data_alignment
            data_style.border = data_border
            
            # Currency style
            currency_font = Font(name='Calibri', size=10)
            currency_number_format = '"₹"#,##0.00'
            currency_style = NamedStyle(name="currency_style")
            currency_style.font = currency_font
            currency_style.number_format = currency_number_format
            currency_style.alignment = Alignment(horizontal='right', vertical='center')
            
            # Register styles
            for style in [title_style, header_style, metric_style, value_style, data_style, currency_style]:
                if style.name not in workbook.named_styles:
                    workbook.add_named_style(style)
            
            # ===== SUMMARY SHEET =====
            ws_summary = workbook.create_sheet("Summary")
            ws_summary.sheet_view.showGridLines = False
            
            # Title
            ws_summary.merge_cells('A1:B1')
            ws_summary['A1'] = "SALES REPORT SUMMARY"
            ws_summary['A1'].style = 'title_style'
            ws_summary.row_dimensions[1].height = 30
            
            # Report period
            ws_summary['A3'] = "Report Period:"
            ws_summary['A3'].style = 'metric_style'
            ws_summary['B3'] = f"{self.start_date.strftime('%d/%m/%Y')} to {self.end_date.strftime('%d/%m/%Y')}"
            ws_summary['B3'].style = 'value_style'
            
            # Generated date
            ws_summary['A4'] = "Generated:"
            ws_summary['A4'].style = 'metric_style'
            ws_summary['B4'] = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
            ws_summary['B4'].style = 'value_style'
            
            # Spacer
            ws_summary['A6'] = "KEY METRICS"
            ws_summary['A6'].style = 'metric_style'
            ws_summary['A6'].font = Font(name='Calibri', size=11, bold=True, color='2C3E50')
            ws_summary.merge_cells('A6:B6')
            ws_summary['A6'].alignment = Alignment(horizontal='center')
            
            # Metrics table
            metrics = [
                ('Total Orders', data['summary']['total_orders']),
                ('Total Revenue', data['summary']['total_revenue']),
                ('Average Order Value', data['summary']['avg_order_value']),
                ('Completed Orders', data['summary']['orders_completed']),
                ('Pending Orders', data['summary']['orders_pending']),
                ('Total Customers', data['summary']['total_customers']),
                ('Report Duration', f"{data['period']['days']} days"),
            ]
            
            for idx, (metric, value) in enumerate(metrics, start=7):
                ws_summary[f'A{idx}'] = metric
                ws_summary[f'A{idx}'].style = 'metric_style'
                
                if 'Revenue' in metric or 'Value' in metric:
                    ws_summary[f'B{idx}'] = float(value)
                    ws_summary[f'B{idx}'].style = 'currency_style'
                else:
                    ws_summary[f'B{idx}'] = value
                    ws_summary[f'B{idx}'].style = 'value_style'
            
            # Set column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            
            # ===== ORDER DETAILS SHEET =====
            if data['orders']:
                ws_orders = workbook.create_sheet("Order Details")
                ws_orders.sheet_view.showGridLines = False
                
                # Title
                ws_orders.merge_cells('A1:N1')
                ws_orders['A1'] = "ORDER DETAILS"
                ws_orders['A1'].style = 'title_style'
                ws_orders.row_dimensions[1].height = 30
                
                # Subtitle
                ws_orders.merge_cells('A2:N2')
                ws_orders['A2'] = f"Showing {len(data['orders'])} orders from {self.start_date.strftime('%d/%m/%Y')} to {self.end_date.strftime('%d/%m/%Y')}"
                ws_orders['A2'].alignment = Alignment(horizontal='center')
                ws_orders['A2'].font = Font(name='Calibri', size=10, color='6B7280')
                
                # Headers
                headers = [
                    'Order Number', 'Date', 'Customer Name', 'Email', 
                    'Product', 'Size', 'Color', 'Quantity', 
                    'Unit Price', 'Total Amount', 'Order Status', 
                    'Payment Status', 'Payment Method'
                ]
                
                for col, header in enumerate(headers, start=1):
                    cell = ws_orders.cell(row=4, column=col)
                    cell.value = header
                    cell.style = 'header_style'
                
                # Data
                for row_idx, order in enumerate(data['orders'], start=5):
                    ws_orders.cell(row=row_idx, column=1, value=order['order_number']).style = 'data_style'
                    ws_orders.cell(row=row_idx, column=2, value=order['created_at'].strftime('%d/%m/%Y %H:%M') if order['created_at'] else '').style = 'data_style'
                    
                    customer_name = f"{order['user__first_name'] or ''} {order['user__last_name'] or ''}".strip()
                    if not customer_name:
                        customer_name = order['user__email'].split('@')[0]
                    ws_orders.cell(row=row_idx, column=3, value=customer_name).style = 'data_style'
                    
                    ws_orders.cell(row=row_idx, column=4, value=order['user__email']).style = 'data_style'
                    ws_orders.cell(row=row_idx, column=5, value=order['product__name'] or 'Unknown').style = 'data_style'
                    ws_orders.cell(row=row_idx, column=6, value=order['variant__sku'] or '').style = 'data_style'
                    ws_orders.cell(row=row_idx, column=7, value='-').style = 'data_style' # Placeholder or remove if redundant
                    ws_orders.cell(row=row_idx, column=8, value=order['quantity']).style = 'data_style'
                    
                    # Currency cells
                    ws_orders.cell(row=row_idx, column=9, value=float(order['unit_price'] or 0)).style = 'currency_style'
                    ws_orders.cell(row=row_idx, column=10, value=float(order['total_amount'] or 0)).style = 'currency_style'
                    
                    ws_orders.cell(row=row_idx, column=11, value=order['order_status'].title()).style = 'data_style'
                    ws_orders.cell(row=row_idx, column=12, value=order['payment_status'].title()).style = 'data_style'
                    
                    payment_method = dict(Order.PAYMENT_METHOD_CHOICES).get(order['payment_method'], order['payment_method'])
                    ws_orders.cell(row=row_idx, column=13, value=payment_method).style = 'data_style'
                    
                    # Alternate row coloring
                    if row_idx % 2 == 0:
                        for col in range(1, 14):
                            ws_orders.cell(row=row_idx, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # Set column widths
                column_widths = [15, 18, 20, 25, 25, 10, 12, 10, 12, 12, 15, 15, 18]
                for col, width in enumerate(column_widths, start=1):
                    ws_orders.column_dimensions[get_column_letter(col)].width = width
                
                # Freeze header row
                ws_orders.freeze_panes = 'A5'
                
                # Add auto filter
                ws_orders.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{len(data['orders']) + 4}"
            
            # ===== PRODUCT PERFORMANCE SHEET =====
            if data['product_sales']:
                ws_products = workbook.create_sheet("Product Performance")
                ws_products.sheet_view.showGridLines = False
                
                # Title
                ws_products.merge_cells('A1:F1')
                ws_products['A1'] = "PRODUCT PERFORMANCE"
                ws_products['A1'].style = 'title_style'
                ws_products.row_dimensions[1].height = 30
                
                # Headers
                product_headers = ['Product Name', 'Brand', 'Category', 'Quantity Sold', 'Total Revenue', 'Average Price']
                for col, header in enumerate(product_headers, start=1):
                    cell = ws_products.cell(row=3, column=col)
                    cell.value = header
                    cell.style = 'header_style'
                
                # Data
                for row_idx, product in enumerate(data['product_sales'], start=4):
                    ws_products.cell(row=row_idx, column=1, value=product['product__name'] or 'Unknown').style = 'data_style'
                    ws_products.cell(row=row_idx, column=2, value=product['product__brand'] or 'Unknown').style = 'data_style'
                    ws_products.cell(row=row_idx, column=3, value=product['product__category__name'] or 'Uncategorized').style = 'data_style'
                    ws_products.cell(row=row_idx, column=4, value=product['quantity_sold']).style = 'data_style'
                    
                    # Currency cells
                    total_rev = float(product['total_revenue'] or 0)
                    ws_products.cell(row=row_idx, column=5, value=total_rev).style = 'currency_style'
                    
                    avg_price = total_rev / product['quantity_sold'] if product['quantity_sold'] > 0 else 0
                    ws_products.cell(row=row_idx, column=6, value=avg_price).style = 'currency_style'
                    
                    # Alternate row coloring
                    if row_idx % 2 == 0:
                        for col in range(1, 7):
                            ws_products.cell(row=row_idx, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # Set column widths
                product_widths = [35, 20, 25, 15, 15, 15]
                for col, width in enumerate(product_widths, start=1):
                    ws_products.column_dimensions[get_column_letter(col)].width = width
                
                # Freeze header row
                ws_products.freeze_panes = 'A4'
                
                # Add auto filter
                ws_products.auto_filter.ref = f"A3:F{len(data['product_sales']) + 3}"
            
            # ===== CATEGORY ANALYSIS SHEET =====
            if data['category_sales']:
                ws_categories = workbook.create_sheet("Category Analysis")
                ws_categories.sheet_view.showGridLines = False
                
                # Title
                ws_categories.merge_cells('A1:C1')
                ws_categories['A1'] = "CATEGORY ANALYSIS"
                ws_categories['A1'].style = 'title_style'
                ws_categories.row_dimensions[1].height = 30
                
                # Headers
                category_headers = ['Category', 'Quantity Sold', 'Total Revenue']
                for col, header in enumerate(category_headers, start=1):
                    cell = ws_categories.cell(row=3, column=col)
                    cell.value = header
                    cell.style = 'header_style'
                
                # Data
                for row_idx, category in enumerate(data['category_sales'], start=4):
                    ws_categories.cell(row=row_idx, column=1, value=category['product__category__name'] or 'Uncategorized').style = 'data_style'
                    ws_categories.cell(row=row_idx, column=2, value=category['quantity_sold']).style = 'data_style'
                    
                    # Currency cell
                    ws_categories.cell(row=row_idx, column=3, value=float(category['total_revenue'] or 0)).style = 'currency_style'
                    
                    # Alternate row coloring
                    if row_idx % 2 == 0:
                        for col in range(1, 4):
                            ws_categories.cell(row=row_idx, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # Set column widths
                category_widths = [30, 15, 15]
                for col, width in enumerate(category_widths, start=1):
                    ws_categories.column_dimensions[get_column_letter(col)].width = width
                
                # Freeze header row
                ws_categories.freeze_panes = 'A4'
                
                # Add auto filter
                ws_categories.auto_filter.ref = f"A3:C{len(data['category_sales']) + 3}"
            
            # ===== DAILY SALES SHEET =====
            if data['daily_sales']:
                ws_daily = workbook.create_sheet("Daily Sales")
                ws_daily.sheet_view.showGridLines = False
                
                # Title
                ws_daily.merge_cells('A1:D1')
                ws_daily['A1'] = "DAILY SALES TREND"
                ws_daily['A1'].style = 'title_style'
                ws_daily.row_dimensions[1].height = 30
                
                # Headers
                daily_headers = ['Date', 'Orders', 'Total Revenue', 'Average Order Value']
                for col, header in enumerate(daily_headers, start=1):
                    cell = ws_daily.cell(row=3, column=col)
                    cell.value = header
                    cell.style = 'header_style'
                
                # Data
                for row_idx, day in enumerate(data['daily_sales'], start=4):
                    ws_daily.cell(row=row_idx, column=1, value=day['date'].strftime('%d/%m/%Y') if day['date'] else '').style = 'data_style'
                    ws_daily.cell(row=row_idx, column=2, value=day['order_count']).style = 'data_style'
                    
                    # Currency cells
                    ws_daily.cell(row=row_idx, column=3, value=float(day['total_revenue'] or 0)).style = 'currency_style'
                    ws_daily.cell(row=row_idx, column=4, value=float(day['avg_order_value'] or 0)).style = 'currency_style'
                    
                    # Alternate row coloring
                    if row_idx % 2 == 0:
                        for col in range(1, 5):
                            ws_daily.cell(row=row_idx, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # Set column widths
                daily_widths = [15, 12, 18, 20]
                for col, width in enumerate(daily_widths, start=1):
                    ws_daily.column_dimensions[get_column_letter(col)].width = width
                
                # Freeze header row
                ws_daily.freeze_panes = 'A4'
                
                # Add auto filter
                ws_daily.auto_filter.ref = f"A3:D{len(data['daily_sales']) + 3}"
            
            # Remove default sheet if exists
            if 'Sheet' in workbook.sheetnames:
                std_sheet = workbook['Sheet']
                workbook.remove(std_sheet)
            
            # Reorder sheets
            sheet_order = ['Summary', 'Order Details', 'Product Performance', 'Category Analysis', 'Daily Sales']
            for sheet_name in reversed(sheet_order):
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    workbook._sheets.insert(0, sheet)
                    workbook._sheets.remove(sheet)
        
        output.seek(0)
        return output.getvalue()
    
    def generate_csv_report(self):
        """Generate comprehensive CSV report"""
        data = self.get_report_data()
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['E-COMMERCE SALES REPORT'])
        writer.writerow([f"Period: {self.start_date.strftime('%d/%m/%Y')} to {self.end_date.strftime('%d/%m/%Y')}"])
        writer.writerow([f"Generated: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([''])
        
        # Summary Section
        writer.writerow(['SUMMARY'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Orders', data['summary']['total_orders']])
        writer.writerow(['Total Revenue', f"Rs.{data['summary']['total_revenue']:,.2f}"])
        writer.writerow(['Average Order Value', f"Rs.{data['summary']['avg_order_value']:,.2f}"])
        writer.writerow(['Completed Orders', data['summary']['orders_completed']])
        writer.writerow(['Pending Orders', data['summary']['orders_pending']])
        writer.writerow(['Total Customers', data['summary']['total_customers']])
        writer.writerow(['Report Duration', f"{data['period']['days']} days"])
        writer.writerow([''])
        
        # Top Products Section
        if data['product_sales']:
            writer.writerow(['TOP PRODUCTS'])
            writer.writerow(['Product', 'Brand', 'Category', 'Quantity', 'Revenue'])
            for product in data['product_sales'][:20]:
                writer.writerow([
                    product['product__name'] or 'Unknown',
                    product['product__brand'] or 'Unknown',
                    product['product__category__name'] or 'Uncategorized',
                    product['quantity_sold'],
                    f"Rs.{float(product['total_revenue'] or 0):,.2f}"
                ])
            writer.writerow([''])
        
        # Order Details Section
        if self.report_type == 'detailed' and data['orders']:
            writer.writerow(['ORDER DETAILS'])
            writer.writerow([
                'Order Number', 'Date', 'Customer Name', 'Email', 'Product',
                'Variant', 'Quantity', 'Unit Price', 'Total Amount', 
                'Order Status', 'Payment Status', 'Payment Method'
            ])
            
            for order in data['orders']:
                customer_name = f"{order['user__first_name'] or ''} {order['user__last_name'] or ''}".strip()
                if not customer_name:
                    customer_name = order['user__email'].split('@')[0]
                
                variant = f"SKU: {order['variant__sku'] or ''}".strip()
                variant = variant if variant else 'Standard'
                
                writer.writerow([
                    order['order_number'],
                    order['created_at'].strftime('%d/%m/%Y %H:%M') if order['created_at'] else '',
                    customer_name,
                    order['user__email'],
                    order['product__name'] or 'Unknown',
                    variant,
                    order['quantity'],
                    f"Rs.{float(order['unit_price']):,.2f}" if order['unit_price'] else "Rs.0.00",
                    f"Rs.{float(order['total_amount']):,.2f}" if order['total_amount'] else "Rs.0.00",
                    order['order_status'].title(),
                    order['payment_status'].title(),
                    dict(Order.PAYMENT_METHOD_CHOICES).get(order['payment_method'], order['payment_method'])
                ])
        
        # Get CSV content
        csv_content = output.getvalue()
        output.close()
        
        return csv_content


def generate_report_response(report_format, start_date=None, end_date=None, report_type='detailed'):
    """Generate HTTP response with report"""
    
    generator = SalesReportGenerator(start_date, end_date, report_type)
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if report_format == 'pdf':
        pdf_content = generator.generate_pdf_report()
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    elif report_format == 'excel':
        excel_content = generator.generate_excel_report()
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    elif report_format == 'csv':
        csv_content = generator.generate_csv_report()
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response['Content-Type'] = 'text/csv; charset=utf-8'
    
    else:
        # Default to PDF if format not specified
        pdf_content = generator.generate_pdf_report()
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    return response